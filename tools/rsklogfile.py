import parse
import json
import argparse
import sys
import ast
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from collections import OrderedDict as odict

parser = argparse.ArgumentParser()
parser.add_argument('logFile', help='log filename')
parser.add_argument('-o', '--output', help='Raw output log operations (XLS format)')
parser.add_argument('-c', '--complete', help='Emit ckpool-like log (Y) or complete event information (default)')
parser.add_argument('-r', '--rskmode', help='Log RSK plugin (Y) or unmodified Stratum Mining (default)')
parser.add_argument('-d', '--debug', help='Debug mode (Y)')
parser.add_argument('-n', '--notify', help='Adds miner notification functionality (Y)')
args = parser.parse_args()

def delta_ms(start, finish):
    delta = finish - start
    return delta * 1000.0

class RSKParser:
    def __init__(self, filename, output, mode, complete, debug):
        self.filename = filename
        self.complete = True if complete == "Y" else False
        self.debug = True if debug == "Y" else False
        self.output = open(output, "w+") if output else None
        self.error_output = open("error.log", "w+") if "error.log" else None
        self.rskmode = True if mode == "Y" else False
        self.rsk_block_received_event = []
        self.rsk_block_received_start = []
        self.rsk_block_received_templ = []
        self.rsk_block_received_end   = []
        self.btc_block_received_event = []
        self.btc_block_received_start = []
        self.btc_block_received_templ = []
        self.btc_block_received_end   = []
        self.share_received_start     = []
        self.share_received_hex       = []
        self.rsk_submitblock          = []
        self.btc_submitblock          = []
        self.gbt_evt_times = []
        self.smblock_times = []
        self.line_count = 1
        self.submitblock_end = []
        self.dumping_conf             = True
        self.confpy                   = []
        self.swb = Workbook()
        self.sws = self.swb.active
        self.sws.title = "rsk-str-merge-mining-summary"
        self.sws.append(["command", "date", "duration1", "id", "clients", "duration2", "last-first"])
        self.sws['A1'].font = Font(bold=True)
        self.sws['B1'].font = Font(bold=True)
        self.sws['C1'].font = Font(bold=True)
        self.sws['D1'].font = Font(bold=True)
        self.sws['E1'].font = Font(bold=True)
        self.sws['F1'].font = Font(bold=True)
        self.sws['G1'].font = Font(bold=True)
        self.sws_rowcount = 2
        if self.complete:
            if self.rskmode:
                if self.debug:
                    self.swb_rbrs = self.swb.create_sheet()
                    self.swb_rbrs.title = "RSK Block Received - Start"
                    self.swb_rbre = self.swb.create_sheet()
                    self.swb_rbre.title = "RSK Block Received - End"
                    self.swb_rbrt = self.swb.create_sheet()
                    self.swb_rbrt.title = "RSK Block Received - Template"
                    self.swb_rsb = self.swb.create_sheet()
                    self.swb_rsb.title = "RSK Submitblock"
                self.swb_rbrev = self.swb.create_sheet()
                self.swb_rbrev.title = "RSK Block Received Event"
                self.swb_rnbp = self.swb.create_sheet()
                self.swb_rnbp.title = "RSK New Block Parent"
                self.swb_rnwu = self.swb.create_sheet()
                self.swb_rnwu.title = "RSK New Work Unit"
            if self.debug:
                self.swb_bbrs = self.swb.create_sheet()
                self.swb_bbrs.title = "BTC Block Received - Start"
                self.swb_bbre = self.swb.create_sheet()
                self.swb_bbre.title = "BTC Block Received - End"
                self.swb_bbrt = self.swb.create_sheet()
                self.swb_bbrt.title = "BTC Block Received - Template"
                self.swb_srs = self.swb.create_sheet()
                self.swb_srs.title = "Share Received - Start"
                self.swb_srh = self.swb.create_sheet()
                self.swb_srh.title = "Share Received - Hex"
                self.swb_sb = self.swb.create_sheet()
                self.swb_sb.title = "Submitblock End"
                self.swb_bsb = self.swb.create_sheet()
                self.swb_bsb.title = "BTC Submitblock"
            self.swb_bbrev = self.swb.create_sheet()
            self.swb_bbrev.title = "BTC Block Received Event"
            self.swb_bnbp = self.swb.create_sheet()
            self.swb_bnbp.title = "BTC New Block Parent"
            self.swb_bnwu = self.swb.create_sheet()
            self.swb_bnwu.title = "BTC New Work Unit"
            self.swb_ws = self.swb.create_sheet()
            self.swb_ws.title = "Work Sent"
            self.swb_wso = self.swb.create_sheet()
            self.swb_wso.title = "Work Sent (Old)"
            self.swb_process = self.swb.create_sheet()
            self.swb_process.title = "CPU MEM %"

        self.notify_events = []
        self.p_notify_events = []
        self.curr_notify_id = ""

    def parse(self):
        with open(self.filename, "r") as f:
            for line in f:
                sys.stdout.write("Parsing line number %d   \r" % self.line_count)
                sys.stdout.flush()
                self.parseline(line)
                self.line_count += 1

        ft = Font(bold=True)

        self.sws['L1'] = "notif pool -> miner"
        self.sws['L1'].font = ft
        self.sws['L2'] = "average"
        self.sws['L3'] = "std dev"
        self.sws['L4'] = "max"
        self.sws['L5'] = "median"
        self.sws['L6'] = "min"
        self.sws['M2'] = '=AVERAGE($I$2:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M3'] = '=STDEV($I$2:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M4'] = '=MAX($I$2:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M5'] = '=MEDIAN($I$2:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M6'] = '=MIN($I$2:$I$' + str(self.sws_rowcount) + ')'

        self.sws['L8'] = "after 30 seconds"
        self.sws['L8'].font = ft
        self.sws['L9'] = "average"
        self.sws['L10'] = "std dev"
        self.sws['L11'] = "max"
        self.sws['L12'] = "median"
        self.sws['L13'] = "min"
        self.sws['M9'] = '=AVERAGE($I$4:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M10'] = '=STDEV($I$4:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M11'] = '=MAX($I$4:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M12'] = '=MEDIAN($I$4:$I$' + str(self.sws_rowcount) + ')'
        self.sws['M13'] = '=MIN($I$4:$I$' + str(self.sws_rowcount) + ')'

        self.sws['O1'] = "notif pool -> miner2"
        self.sws['O1'].font = ft
        self.sws['O2'] = "average"
        self.sws['O3'] = "std dev"
        self.sws['O4'] = "max"
        self.sws['O5'] = "median"
        self.sws['O6'] = "min"
        self.sws['P2'] = '=AVERAGE($J$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P3'] = '=STDEV($J$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P4'] = '=MAX($J$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P5'] = '=MEDIAN($J$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P6'] = '=MIN($J$2:$J$' + str(self.sws_rowcount) + ')'

        self.sws['O8'] = "after 30 seconds"
        self.sws['O8'].font = ft
        self.sws['O9'] = "average"
        self.sws['O10'] = "std dev"
        self.sws['O11'] = "max"
        self.sws['O12'] = "median"
        self.sws['O13'] = "min"
        self.sws['P9'] = '=AVERAGE($J$4:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P10'] = '=STDEV($J$4:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P11'] = '=MAX($J$4:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P12'] = '=MEDIAN($J$4:$J$' + str(self.sws_rowcount) + ')'
        self.sws['P13'] = '=MIN($J$4:$J$' + str(self.sws_rowcount) + ')'

        self.sws['S1'] = "submit miner -> bitcoin"
        self.sws['S1'].font = ft
        self.sws['S2'] = "average"
        self.sws['S3'] = "std dev"
        self.sws['S4'] = "max"
        self.sws['S5'] = "median"
        self.sws['S6'] = "min"
        self.sws['T2'] = '=AVERAGE($Q$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['T3'] = '=STDEV($Q$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['T4'] = '=MAX($Q$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['T5'] = '=MEDIAN($Q$2:$J$' + str(self.sws_rowcount) + ')'
        self.sws['T6'] = '=MIN($Q$2:$J$' + str(self.sws_rowcount) + ')'

        self.sws.column_dimensions.group('I','J', hidden=True)
        self.sws.column_dimensions.group('Q', hidden=True)
        print "File parsed and loaded in memory"

        if self.output:
            self.swb.save(self.output)

    def parseline(self, line):
        if self.dumping_conf:
            if line.find("END CONFIG.PY DUMP") > 0:
                self.dumping_conf = False
                fo = open("config.py", "wb")
                for i in self.confpy:
                    fo.write(i)

                fo.close()
                return
            else:
                ln = line.split("mining")
                if len(ln) > 1:
                    self.confpy.append(ln[1])

        if line.find("RSKLOG") < 0 and line.find("STRLOG") < 0 and line.find("ACCEPTED") < 0 and line.find("REJECTED") < 0 and line.find("MINNOT"):
            return

        if line.find("MINER EMIT") > 0:
            return

        if line.find("MINNOT") > 0:
            res = ast.literal_eval(line.split('#')[1].strip())
            event = res['tag']
        elif line.find("ACCEPTED") > 0:
            block = (line.split('#')[1].split(' ')[2], str(line.split('#')[0].split(' ')[0] + " " + line.split('#')[0].split(' ')[1]))
            self.process_submitblock_event(block)
            return
        elif line.find("REJECTED") > 0:
            block = line.split('#')[1].split(' ')[2]
            self.submitblock_end = [x for x in self.submitblock_end if x['data'][0] != block]
            self.btc_submitblock = [x for x in self.btc_submitblock if x['data'] != block]
            if self.rskmode:
                self.rsk_submitblock = [x for x in self.rsk_submitblock if x['data'] != block]
            return
        else:
            res = json.loads(line.split('#')[1].strip())
            event = res['tag']

        if event == "[MINNOT]":
            if self.curr_notify_id != "":
                if res['data'] != self.curr_notify_id:
                    n_ev = {"start" : self.timestamp_to_str(self.notify_events[0]['start']), "delta_ms" : delta_ms(self.notify_events[0]['start'], self.notify_events[-1]['start']), "id" : res['data']}
                    self.p_notify_events.append(n_ev)
                    self.notify_events = []
                    self.curr_notify_id = res['data']
                    self.notify_events.append(res)
                    brt_match = [ x for x in self.btc_block_received_templ if x['data'][0] == self.curr_notify_id]
                    if len(brt_match) > 0:
                        self.process_btc_blockreceived_event(brt_match)
                else:
                    self.notify_events.append(res)
            else:
                self.curr_notify_id = res['data']
                self.notify_events.append(res)

        if self.rskmode:
            if self.complete:
                if event == "[RSK_NEW_BLOCK_PARENT]":
                    self.swb_rnbp.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
                elif event == "[RSK_NEW_WORK_UNIT]":
                    self.swb_rnwu.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
            if event == "[RSK_BLOCK_RECEIVED_START]":
                if self.debug:
                    self.swb_rbrs.append([res['uuid'], res['start']])
                self.rsk_block_received_start.append(res)
            elif event == "[RSK_BLOCK_RECEIVED_TEMPLATE]":
                if self.debug:
                    self.swb_rbrt.append([res['uuid'], res['start'], res['data']])
            elif event == "[RSK_BLOCK_RECEIVED_END]":
                if self.debug:
                    self.swb_rbre.append([res['uuid'], res['start'], res['data']])
                self.rsk_block_received_end.append(res)
            elif event == "[RSK_SUBMITBLOCK]":
                if self.debug:
                    self.swb_rsb.append([res['uuid'], res['start'], res['data']])
                self.rsk_submitblock.append(res)

        if self.complete:
            if event == "[BTC_NEW_BLOCK_PARENT]":
                self.swb_bnbp.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
            elif event == "[BTC_NEW_WORK_UNIT]":
                self.swb_bnwu.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
            elif event == "[WORK_SENT]":
                self.swb_ws.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
            elif event == "[WORK_SENT_OLD]":
                self.swb_wso.append([res['uuid'], self.timestamp_to_str(res['start']), (float(res['elapsed'] * 1000))])
            elif event == "[PROCESS]":
                self.swb_process.append([self.timestamp_to_str(res['start']), res['data']['threads'][0]['cpu_percent'], res['data']['threads'][1]['cpu_percent'], res['data']['memory_percent']])

        if event == "[BTC_BLOCK_RECEIVED_START]":
            if self.debug:
                self.swb_bbrs.append([res['uuid'], res['start']])
            self.btc_block_received_start.append(res)
        elif event == "[BTC_BLOCK_RECEIVED_TEMPLATE]":
            if self.debug:
                self.swb_bbrt.append([res['uuid'], res['start'], res['data']])
            self.btc_block_received_templ.append(res)
            #self.process_btc_blockreceived_event(res)
        elif event == "[BTC_BLOCK_RECEIVED_END]":
            if self.debug:
                self.swb_bbre.append([res['uuid'], res['start'], res['data']])
            self.btc_block_received_end.append(res)
        elif event == "[SHARE_RECEIVED_START]":
            if self.debug:
                self.swb_srs.append([res['uuid'], res['start']])
            self.share_received_start.append(res)
        elif event == "[SHARE_RECEIVED_HEX]":
            if self.debug:
                self.swb_srh.append([res['uuid'], res['start'], res['data']])
            self.share_received_hex.append(res)
        elif event == "[BTC_SUBMITBLOCK]":
            if self.debug:
                self.swb_bsb.append([res['uuid'], res['start'], res['data']])
            self.btc_submitblock.append(res)
        elif event == "[SUBMITBLOCK_END]":
            if self.debug:
                self.swb_sb.append([res['uuid'], res['start'], res['data']])
            self.submitblock_end.append(res)

    def timestamp_to_str(self, tm):
        return datetime.fromtimestamp(tm).strftime('%Y-%m-%d %H:%M:%S.%f')

    def process_btc_blockreceived_event(self, ev):
        end_match = [x for x in self.btc_block_received_end if x['data'] == ev[0]['data']][0]
        sta_match = [x for x in self.btc_block_received_start if x['uuid'] == ev[0]['uuid']][0]
        not_match = [x for x in self.p_notify_events if x['id'] == end_match['data'][0]][0]

        dat = {"uuid" : sta_match['uuid'], "start" : self.timestamp_to_str(sta_match['start']), "delta_gbt" : delta_ms(sta_match['start'], end_match['start']), "delta_emit" : delta_ms(ev[0]['start'], float(end_match['start'] + end_match['elapsed'])), "clients" : end_match['clients']}
        if self.complete:
            self.swb_bbrev.append([dat['uuid'], dat['start'], dat['delta_gbt'], dat['clients'], dat['delta_emit'], not_match["delta_ms"]])

        self.sws.append(["getblocktemplate", dat['start'], dat['delta_gbt'], dat['uuid'], dat['clients'], (dat['delta_emit'] + not_match["delta_ms"]), float(not_match["delta_ms"]), '',
            '=IF(A' + str(self.sws_rowcount) + '=\"getblocktemplate\", F' + str(self.sws_rowcount) +', \"\")',
            '=IF(A' + str(self.sws_rowcount) + '=\"getblocktemplate\", G' + str(self.sws_rowcount) +', \"\")'])
        self.sws_rowcount += 1
        self.gbt_evt_times.append([dat['delta_gbt'], dat['delta_emit'], not_match['delta_ms']])
        self.btc_block_received_end = [x for x in self.btc_block_received_end if x['uuid'] != end_match['uuid']]
        self.btc_block_received_start = [x for x in self.btc_block_received_start if x['uuid'] != sta_match['uuid']]
        self.btc_block_received_templ = [x for x in self.btc_block_received_templ if x['uuid'] != ev[0]['uuid']]

    def str_to_timestamp(self, s):
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S,%f")
        return time.mktime(dt.timetuple()) + (dt.microsecond / 1000000.0)

    def sb_match_helper(self, match, ev, acc):
        n_ev = {"start" : "", "delta_process" : "", "delta_emit" : "", "hex" : ""}
        shhex_match = [x for x in self.share_received_hex if x['data'] == ev['data'][0]][0]
        shstr_match = [x for x in self.share_received_start if x['uuid'] == shhex_match['uuid']][0]
        n_ev["start"] = self.timestamp_to_str(shstr_match["start"])
        n_ev["delta_process"] = delta_ms(ev["start"], shhex_match["start"])
        n_ev["delta_emit"] = delta_ms(shhex_match["start"], match["elapsed"])
        n_ev["hex"] = ev["data"][0]

        self.share_received_hex = [x for x in self.share_received_hex if x['start'] < (ev['start'] - 60)]
        self.share_received_start = [x for x in self.share_received_start if x['start'] < (ev['start'] - 60)]
        self.submitblock_end = [x for x in self.submitblock_end if x['start'] < (ev['start'] - 60)]
        self.rsk_submitblock = [x for x in self.rsk_submitblock if x['start'] < (ev['start'] - 60)]
        self.btc_submitblock = [x for x in self.btc_submitblock if x['start'] < (ev['start'] - 60)]

        return n_ev

    def process_submitblock_event(self, ev):
        sbend_match = [x for x in self.submitblock_end if x['data'][0] == ev[0]][0]
        btcsb_match = [x for x in self.btc_submitblock if x['data'] == sbend_match['data'][0]]
        not_match = [x for x in self.p_notify_events if x['id'] == sbend_match['data'][1]]
        if len(btcsb_match) > 0 and len(not_match) > 0:
            dat = self.sb_match_helper(btcsb_match[0], sbend_match, ev)
            self.sws.append(["submitblock", dat['start'], dat['delta_emit'], sbend_match['data'][1], '-', dat['delta_process'], '-', '',
                '=IF(A' + str(self.sws_rowcount) + '=\"getblocktemplate\", F' + str(self.sws_rowcount) +', \"\")',
                '=IF(A' + str(self.sws_rowcount) + '=\"getblocktemplate\", G' + str(self.sws_rowcount) +', \"\")'])
            qstr = 'Q' + str(self.sws_rowcount)
            self.sws[qstr] = dat['delta_emit']
            self.sws_rowcount += 1
            self.btc_submitblock = [x for x in self.btc_submitblock if x['data'] != sbend_match['data']]

def main():
    logfile = RSKParser(args.logFile, args.output, args.rskmode, args.complete, args.debug)
    logfile.parse()

if __name__ == "__main__":
    main()
